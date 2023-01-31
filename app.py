from flask import Flask, request, render_template,Response,redirect,url_for
from werkzeug.utils import secure_filename
import openpyxl,smtplib,sys,imaplib,email,pprint,os,re,PyPDF2,docx,datetime
import numpy as np 
import pandas as pd 
import string
import os
from sys import argv
import docx2txt
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from nltk.stem.porter import PorterStemmer
from nltk.stem import WordNetLemmatizer

from operator import itemgetter
import nltk
nltk.download('stopwords')
stopwords = nltk.corpus.stopwords.words('english')
porter_stemmer = PorterStemmer()
wordnet_lemmatizer = WordNetLemmatizer()




# clean function 
def clean_resume(text):
    text = text.lower() 
    text = ''.join([word for word in text if not word.isdigit()]) 
    text = re.sub('http\S+\s*', ' ', text)  
    # remove RT and cc
    text = re.sub('RT|cc', ' ', text)
    text = re.sub('#\S+', '', text)  
    text = re.sub('@\S+', '  ', text)  
    text = "".join([word for word in text if word not in string.punctuation])
    text = re.sub("\W", " ", str(text))
    ext = [word for word in text.split() if word not in stopwords]
    #replace consecutive non-ASCII characters with a space
    text = re.sub(r'[^\x00-\x7f]',r' ', text) 
    #extra whitespace removal
    text = re.sub('\s+', ' ', text)
    return text

# clean jd function 
def clean_jd(text):
    text = text.lower() 
    text = re.sub(r'[0-9]+', '', text)
    text = re.sub('http\S+\s*', ' ', text)  
    # remove RT and cc
    text = re.sub('RT|cc', ' ', text)
    text = re.sub('#\S+', '', text)  
    text = re.sub('@\S+', '  ', text)  
    text = re.sub("\W", " ", str(text))
    #replace consecutive non-ASCII characters with a space
    text = re.sub(r'[^\x00-\x7f]',r' ', text) 
    #extra whitespace removal
    text = re.sub('\s+', ' ', text)
    return text


app = Flask(__name__)
pat = re.compile(r'\s+')
pat1=re.compile(r'\n+')
received_from_data={}

detach_dir = 'resumes-and-candidate-data\\'
now_time=str(datetime.datetime.now())
micro_second_index=now_time.index('.')
now_time=now_time[:micro_second_index]
detach_dir=detach_dir+now_time # name the new folder with now time 
detach_dir=detach_dir.replace(' ',',')
detach_dir=detach_dir.replace(':','-')

if not os.path.exists(detach_dir):
    os.makedirs(detach_dir)
    
list=[]
result=[]


def inboxSearch(useremail,Ema_password):
    #print('Searching for the resumes...\n\n')
    m = imaplib.IMAP4_SSL("imap.gmail.com")
    try:
        m.login(useremail,Ema_password)
        m.select("inbox")

    except imaplib.IMAP4.error as e:
        if 'Invalid credentials' in str(e):
            print("It seems that password was incorrect.")
    else:
        resp, items = m.search(None,'(UNSEEN SUBJECT "resume")',)
        items = items[0].split()
        
        while(len(items)>0):
            try:
                emailid=items[len(items)-1]
                resp, data = m.fetch(emailid, "(RFC822)") ##fetch email from gived email id  ##defines an electronic message format consisting of header fields and an optional message body. 
                email_body = data[0][1]#store the email id and body 
                email_body=email_body.decode('utf-8')#to convert it from binary 
                mail = email.message_from_string(email_body) #convert the string content as object.
                
                temp = m.store(emailid,'+FLAGS', '\\Seen')#change the massage as seen 
                m.expunge()#delete the msg from the server 
                items.pop()#dont use this item again 

                if mail.get_content_maintype() != 'multipart':#if contain is attachment 
                    continue

                received_from=mail["From"]#save the sender email from(name)
                email_start_index=received_from.index('<')+1
                email_end_index=received_from.index('>')
                received_from_emailid=received_from[email_start_index:email_end_index]
                received_from_name=received_from[:email_start_index-1]
                received_from_date=mail["Date"]#save the date 
            
                #print ("["+mail["From"]+"] :" + mail["Subject"]) # print the candinate email name and subject 

                for part in mail.walk(): #iterate inside the message object 
                    if part.get_content_maintype() == 'multipart':## if contain attachment 
                        continue
                    if part.get('Content-Disposition') is None:# attachment but inline 
                        continue
                    if part.get_filename().endswith('.pdf'):#Check the file type 
                        file_type='.pdf'
                    if part.get_filename().endswith('.docx'):
                        file_type='.docx'
                
                ## Save the resumes locally 
                    filename = received_from_emailid+file_type# define the name of the attachment when saved localy 
                    att_path = os.path.join(detach_dir, filename)

                    if not os.path.isfile(att_path):#if path dose not exsist
                        fp = open(att_path, 'wb')
                        fp.write(part.get_payload(decode=True))
                        fp.close()
                        received_from_data[received_from_emailid]=[received_from_emailid,received_from_name,received_from_date,att_path]
                        print(received_from_data)
            except Exception as r:
                print(r)
        return received_from_data

def extractText(j_description):
    result_list=[]
    print('Scanning all the resumes...\n\n')
    for downloaded_resume in received_from_data:
        content=''
        if received_from_data[downloaded_resume][3].endswith('.pdf'):
            path=received_from_data[downloaded_resume][3]
            pdfFileObj = open(path, 'rb')
            pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
            content=''
            pages=pdfReader.numPages
            
            for i in range(pages):
                pageObj = pdfReader.getPage(i)
                content+=pageObj.extractText()#extract from the specifc page 
            cleaned_res=clean_resume(content)
            cleanedjs=clean_jd(j_description)
            doc=[cleaned_res,cleanedjs]
            #cv = CountVectorizer(ngram_range=(2,2))
            cv = CountVectorizer()
            count_matrix1 = cv.fit_transform(doc)
            similarPercentage1 = cosine_similarity(count_matrix1)[0][1] * 100
            similarPercentage1 = round(similarPercentage1, 2)
            print("("+str(received_from_data[downloaded_resume][0])+ " resume) similiar about "+str(similarPercentage1)+ "% of the job description.")
            resume_name=received_from_data[downloaded_resume][0]


        if received_from_data[downloaded_resume][3].endswith('.docx'):
            path=received_from_data[downloaded_resume][3]
            doc = docx.Document(path)
            content=[]
            for para in doc.paragraphs:
                content.append(para.text)
            content=''.join(para.text)
            cleaned_res=clean_resume(content)
            cleanedjs=clean_jd(j_description)
            list=[cleaned_res,cleanedjs]
            cv = CountVectorizer()
            count_matrix1 = cv.fit_transform(list)
            similarPercentage1 = cosine_similarity(count_matrix1)[0][1] * 100
            similarPercentage1 = round(similarPercentage1, 2) # round to two decimal
            print("("+str(received_from_data[downloaded_resume][0])+ " resume) similiar about "+str(similarPercentage1)+ "% of the job description.")
            resume_name=received_from_data[downloaded_resume][0]

            
            
        if (similarPercentage1 >=40):
            
            status="Pass"
            result_list.append(resume_name)
            result_list.append(similarPercentage1)
            result_list.append(status)
            
        else:
            status="Reject"
            result_list.append(resume_name)
            result_list.append(similarPercentage1)
            result_list.append(status)
        
        if (similarPercentage1 >=40):
            received_from_data[downloaded_resume].append('Yes')#
        else:
            received_from_data[downloaded_resume].append('No')
        
        

    #print(received_from_data)
    print('Finished scanning all the resumes.\n\n')
    return result_list

def saveInXl(received_from_data):
    #received_from_data[id]=[id,name,date,filepath,decision]
    print('Saving data in excel sheet...\n')
    wb=openpyxl.Workbook()
    sheet=wb.active
    sheet.title='resumes'
    sheet.cell(row=1,column=1).value='NAME'
    sheet.cell(row=1,column=2).value='EMAIL ID'
    sheet.cell(row=1,column=3).value='DATE-TIME'
    sheet.cell(row=1,column=4).value='DECISION'

    sheet_row=2
    for downloaded_resume in received_from_data:
        sheet.cell(row=sheet_row,column=1).value=received_from_data[downloaded_resume][1]
        sheet.cell(row=sheet_row,column=2).value=received_from_data[downloaded_resume][0]
        sheet.cell(row=sheet_row,column=3).value=received_from_data[downloaded_resume][2]
        sheet.cell(row=sheet_row,column=4).value=received_from_data[downloaded_resume][4]
        sheet_row+=1


    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 20
    wb.save(detach_dir+ '\\candidate_data.xlsx')
    print("Finished saving data in excel sheet.\n\n")

def sendmail(email,Ema_password):
    state=[]
    print("Sending replies to candidates...\n ")
    smtpObj = smtplib.SMTP('smtp.gmail.com', 587)# Make a Secure connection 
    smtpObj.ehlo()
    smtpObj.starttls()
    useremail=email
    password=Ema_password
    smtpObj.login(useremail,password)
    wb=openpyxl.load_workbook(detach_dir+ '\\candidate_data.xlsx')
    sheet=wb['resumes']
    lastCol=4

    for r in range(2, 2+len(received_from_data)):
        decision=sheet.cell(row=r, column=lastCol).value
        name=sheet.cell(row=r, column=1).value
        senderemail=sheet.cell(row=r, column=2).value
        if decision=='Yes':
            body = "Subject: SELECTED.\nDear %s,\n We are glad to inform you that you are selected for the job interview." %(name)

        else:
            body = "Subject: rejected.\nDear %s,\n We are sorry to inform you that you are not selected for the job interview." %(name)
        
        print('Sending email to %s...' % senderemail)
        sendmailStatus = smtpObj.sendmail(useremail, senderemail, body)

        if sendmailStatus != {}:
            print('There was a problem sending email to %s: %s' % (senderemail,sendmailStatus))


    smtpObj.quit()
    print("Finished sending replies to candidates.\n\n")
            


@app.route('/')
def show():
   return render_template('form.html')

@app.route('/sucsess',methods=['GET','POST'])
def sucess():
    return render_template('RecommendResult.html')

@app.route('/result',methods=['GET','POST'])
def my_link():
    print("my link fuction start ")
    if request.method == 'POST':
      useremail= request.form.get("email")
      Ema_password = request.form.get("password") 
      f = request.files['file']
      j_description = docx2txt.process(f)
      recived_from_data=inboxSearch(useremail,Ema_password)
      inboxSearch(useremail,Ema_password)
      cos_result=extractText(j_description)
      saveInXl(recived_from_data)
      sendmail(useremail,Ema_password)
      #print(cos_result)
      return render_template("RecommendResult.html",cos_result=cos_result)
    
    return render_template("result.html")

if __name__ == "__main__":
    app.run(Debug=True)
